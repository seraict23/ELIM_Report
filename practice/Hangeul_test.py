import os
from tkinter.filedialog import askopenfilenames

import win32com.client as win32   
# Module for opening hangeul

def Hangeul_Start():
    hwp=win32.gencache.EnsureDispatch("HWPFrame.HwpObject") 
    # Start Hangeul
    hwp.XHwpWindows.Item(0).Visible=True
    # Disable background mode
    hwp.RegisterModule("FilePathCheckDLL","SecurityModule")   
    # Register module
    
    return hwp

def SelectFiles():
    filelist=askopenfilenames(title="Select Files",
                             initialdir=os.getcwd(),
                             filetypes=[("한/글 파일","*.hwp *.hwpx")])
    
    return filelist

if __name__=='__main__':
    hwp=Hangeul_Start()
    # Start Hangeul
    
    filelist=SelectFiles()
    for file in filelist:
        if file.endswith("x"):
            extension="hwpx"
        else:
             extension="hwp"
        hwp.Open(file,Format=extension.upper(),arg="")       

## Exercise 1. 상장 만들기
# Ctrl+k+e: 누름틀 만들기
import win32com.client as win32
import pandas as pd 

hwp=win32.Dispatch("HWPFrame.HwpObject")
hwp.XHwpWindows.Item(0).Visible=True
hwp.RegisterModule("FilePathCheckDLL","SecurityModule")   
hwp.Open(r'C:\Users\k\Downloads\certificate_of_award\h25_certificate_of_award.hwp','','')
# Need full path to open haungeul documents

excel=pd.read_excel(r'C:\Users\k\Downloads\certificate_of_award\Award.xlsx')
excel

print(hwp.GetFieldList(0,0))
# Check the field names

fieldlist=[i for i in hwp.GetFieldList(0,0).split('\x02')]

hwp.Run('SelectAll')
hwp.Run('Copy')
hwp.MovePos(3,0,0)
## Move the cursor to the bottom of the document

for i in range(len(excel)-1):
    hwp.Run('Paste')
    hwp.MovePos(3,0,0)
    
for page in range(len(excel)):
    for field in fieldlist:
        hwp.PutFieldText(f'{field}{{{{{page}}}}}',excel[field].iloc[page])

## Hide Hangeul window
import win32gui

# hwp=win32.Dispatch('HWPFrame.HwpObject')
hwnd=win32gui.FindWindow(None,'빈 문서 1 - 한글')
win32gui.ShowWindow(hwnd,5)

## Exercise 2. Convert HWP to PDF (Multiple)
os.chdir(r'C:\Users\k\Downloads\HWP to PDF')
# Change directory

for i in os.listdir():
    os.rename(i,i.replace(' - Copy',''))

hwp=win32.Dispatch("HWPFrame.HwpObject")
hwp.XHwpWindows.Item(0).Visible=True
hwp.RegisterModule("FilePathCheckDLL","SecurityModule")

base_dir=r'C:\Users\k\Downloads\HWP to PDF'

os.chdir(base_dir)

for i in os.listdir():
    hwp.Open(os.path.join(base_dir,i),'','')
    hwp.HAction.GetDefault('FileSaveAsPdf',hwp.HParameterSet.HFileOpenSave.HSet)
    hwp.HParameterSet.HFileOpenSave.filename=os.path.join(base_dir,i.replace('.hwp','.pdf'))
    hwp.HParameterSet.HFileOpenSave.Format='PDF'
    hwp.HAction.Execute("FileSaveAsPdf",hwp.HParameterSet.HFileOpenSave.HSet)
# Check when initial dialog box pops up    

## Exercise 3. 금주실적표 작성
# Ctrl+n+k 셀 탭에 필드 입력하여 표용 누름틀 생성
os.chdir(r"C:\Users\k\Documents\Hangeul Sample\weekly_report")

os.listdir()

import shutil
# used when copy, paste, or move files
import re
# regualr expression

file=os.listdir()[0]
for i in range(2,53):
    shutil.copy(file,file.split('(')[0]+f'({i:0>2}주차).hwp')

hwp=win32.Dispatch("HWPFrame.HwpObject")
hwp.RegisterModule("FilePathCheckDLL","SecurityModule")
# Minimize window or turn on background mode

from time import sleep
from random import randint
import datetime as dt

date=dt.datetime(2023,1,1)
delta=dt.timedelta(days=1)

workSum=0
for i in os.listdir():
#     To eliminate 0 at the first place(int)
    hwp.Open(os.path.join(os.getcwd(),i),'','')
    sleep(0.5)
    
    weekNum=int(re.findall(r'\d+',i)[1])
    hwp.PutFieldText('WeekNum',weekNum)
    
    date+=delta
    hwp.PutFieldText('Mon',date.strftime("%Y. %m. %d.(월)"))
#     Convert datetime to formatted string (strftime)
    hwp.PutFieldText('MonWork',randint(1,10))
    workSum+=int(hwp.GetFieldText('MonWork'))
    hwp.PutFieldText('MonWorkSum',workSum)
    
    date+=delta
    hwp.PutFieldText('Tue',date.strftime("%Y. %m. %d.(화)"))
    hwp.PutFieldText('TueWork',randint(1,10))
    workSum+=int(hwp.GetFieldText('TueWork'))
    hwp.PutFieldText('TueWorkSum',workSum)
    
    date+=delta
    hwp.PutFieldText('Wed',date.strftime("%Y. %m. %d.(수)"))
    hwp.PutFieldText('WedWork',randint(1,10))
    workSum+=int(hwp.GetFieldText('WedWork'))
    hwp.PutFieldText('WedWorkSum',workSum)
    
    date+=delta
    hwp.PutFieldText('Thur',date.strftime("%Y. %m. %d.(목)"))
    hwp.PutFieldText('ThurWork',randint(1,10))
    workSum+=int(hwp.GetFieldText('ThurWork'))
    hwp.PutFieldText('ThurWorkSum',workSum)
    
    date+=delta
    hwp.PutFieldText('Fri',date.strftime("%Y. %m. %d.(금)"))
    hwp.PutFieldText('FriWork',randint(1,10))
    workSum+=int(hwp.GetFieldText('FriWork'))
    hwp.PutFieldText('FriWorkSum',workSum)
    
    date+=delta
    hwp.PutFieldText('Sat',date.strftime("%Y. %m. %d.(토)"))
    hwp.PutFieldText('SatWork','-')
    hwp.PutFieldText('SatWorkSum','-')
    
    date+=delta
    hwp.PutFieldText('Sun',date.strftime("%Y. %m. %d.(일)"))
    hwp.PutFieldText('SunWork','-')
    hwp.PutFieldText('SunWorkSum','-')
        
    hwp.Save(True)
    sleep(0.5)
    # Prevent creating savefile dialog
    
    hwp.Run('FileClose')

## Exercise 4. Make list page without outline
# 도구>차례/색인>차례 만들기 이용
# 개요는 목차에 사용, 문단 번호는 그 외 번호 매기는 데 사용

hwp=win32.Dispatch("HWPFrame.HwpObject")
hwp.XHwpWindows.Item(0).Visible=True
hwp.RegisterModule("FilePathCheckDLL","SecurityModule")
hwp.Open(r"C:\Users\k\Documents\Hangeul Sample\make_list\Lorem Ipsum.hwp",'','')

Poslist=[]
hwp.InitScan(None,None,None,None,None,None)
while True:
    text=hwp.GetText()
    print(text)
    if text[0]==0:    # [0]=status, [1]=content, When it reaches the end of the document(0)
        hwp.ReleaseScan()
        break
    else:
        if re.match(r'[\d+가-힣]\.',text[1].strip()):   # Depends on document format
            hwp.MovePos(201,None,None)
            Poslist.append(hwp.GetPos())
            
for Pos in Poslist:
    hwp.SetPos(*Pos)    # Convert tuple into multiple int variables(*)
    hwp.Run("MarkTitle")    # Insert [제목차례]

hwp.MovePos(2,None,None)    # moveTopOfFile
hwp.Run('BreakPage')    # Make list page
hwp.MovePos(2,None,None)

hwp.HAction.GetDefault("MakeContents",hwp.HParameterSet.HMakeContents.Hset)
hwp.HParameterSet.HMakeContents.Make=65    # 0x01(1)+0x40(64)=65
hwp.HParameterSet.HMakeContents.Leader=6
hwp.HParameterSet.HMakeContents.type=1
hwp.HAction.Execute("MakeContents",hwp.HParameterSet.HMakeContents.HSet)

## Exercise 5-1. Webcrawling then make a report
from time import sleep

import pyperclip as cb
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
import win32com.client as win32
import lxml

driver=webdriver.Chrome(r"./chromedriver_win32/chromedriver.exe")
driver.get("https://dart.fss.or.kr/dsac001/mainAll.do")

onclick_list=[
    i.get_attribute('onclick') for i in driver.find_elements(By.CSS_SELECTOR,'a[href^="/dsaf001/main.do?rcpNo="]')]
# Much faster when opening javascript directly than clicking
# ^: begins with (regex)

link_list=[]
for i in onclick_list:
    driver.execute_script(i)
    driver.switch_to.window(driver.window_handles[1])
    driver.execute_script(driver.find_element(By.CSS_SELECTOR,'.btnDown').get_attribute('onclick'))
    driver.switch_to.window(driver.window_handles[2])
    current_len=len(link_list)
    link_list.append(driver.find_element(By.CSS_SELECTOR,'a.btnFile').get_attribute('href'))

    while True:
        if(current_len!=len(link_list)):
            break
        else:
            sleep(0.1)
    # prevent closing window before getting attribute

    driver.close()
    driver.switch_to.window(driver.window_handles[1])
    driver.close()
    driver.switch_to.window(driver.window_handles[0])

table=pd.read_html(driver.page_source)[0]

driver.close()

hwp=win32.Dispatch("HWPFrame.HwpObject")
hwp.XHwpWindows.Item(0).Visible=True
hwp.RegisterModule("FilePathCheckDLL","SecurityModule")
hwp.Open(r"C:\Users\k\Documents\Hangeul Sample\webcrawling\dart_template.hwp",'','')
# Cursor must be in cell A2 of the table

def hwp_insert_hyperlink(text,url):
    hwp.HAction.Run("TableCellBlock")
    hwp.HAction.GetDefault("InsertHyperlink",hwp.HParameterSet.HHyperLink.HSet)
    hwp.HParameterSet.HHyperLink.Text=text
    hwp.HParameterSet.HHyperLink.Command=url
    hwp.HAction.Execute("InsertHyperlink",hwp.HParameterSet.HHYperLink.HSet)

def hwp_insert_text(text):
    hwp.HAction.GetDefault("InsertText",hwp.HParameterSet.HInsertText.HSet)
    hwp.HParameterSet.HInsertText.Text=text
    hwp.HAction.Execute("InsertText",hwp.HParameterSet.HInsertText.HSet)

for i in range(len(table)):
    for j,text in enumerate(table.loc[i]):
        if i==0 and j==0:            
            hwp_insert_text(text)
        elif j==2:
            hwp.Run("TableRightCellAppend")
            hwp_insert_hyperlink(text,'{};1;0;0'.format(link_list[i].replace("?","\\?")))
            # Without \\, ? is accepted as a special symbol
        elif j==5 and text=="1.#QNAN":
            hwp.Run("TableRightCellAppend")
        else:
            hwp.Run("TableRightCellAppend")
            hwp_insert_text(text)

## Exercise 5-2. Webcrawling then make a report
import os
from time import sleep
import requests
# Download images
from io import BytesIO
# Copy image to clipboard
import shutil

import pandas as pd
import win32clipboard
import win32com.client as win32
from PIL import Image
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.support.select import Select
# Deal with exceptions when webcrawling
from selenium.webdriver.common.by import By

def Execute_script(script):
    while True:
        try:
            driver.execute_script(script)
            break
        except JavascriptException:
            sleep(0.5)
# Prevent javascript exceptions

download_dir=r"C:\Users\k\Documents\Hangeul Sample\webcrawling_2\Imgs"

driver=webdriver.Chrome('./chromedriver_win32/chromedriver.exe')
driver.get("https://www.g2b.go.kr:8092/sm/ma/mn/SMMAMnF.do")
driver.switch_to.frame('sub')

driver.find_element(By.CSS_SELECTOR,'#kwd.srch_txt').send_keys('작업용의자')
driver.find_element(By.CSS_SELECTOR,'#kwd.srch_txt').submit()

Execute_script(driver.find_element(By.CSS_SELECTOR,'div#btnAttrDivHide.btnAttr>a').get_attribute("href"))
Execute_script(driver.find_element(By.CSS_SELECTOR,'#attrDtl01Div > li:nth-child(7) > a').get_attribute("href"))
sleep(1)

checklist_PP=driver.find_elements(By.CSS_SELECTOR,'[id^="chkAttrVal"][value*="플라스틱"]')
checklist_메쉬=driver.find_elements(By.CSS_SELECTOR,'[id^="chkAttrVal"][value*="메쉬"]')
checklist_가죽=driver.find_elements(By.CSS_SELECTOR,'[id^="chkAttrVal"][value*="가죽"]')
# *=: contains, ^=begins with, $=ends with (CSS selector)

for check_메쉬 in checklist_메쉬:
    if not check_메쉬 in checklist_가죽:
        check_메쉬.click()       

for check_PP in checklist_PP:
    if not check_PP in checklist_가죽:
        check_PP.click()        

Execute_script(driver.find_element(By.CSS_SELECTOR,'#btnAttrInit > a:nth-child(2)').get_attribute("onclick"))

Execute_script(driver.find_element(By.CSS_SELECTOR,'#attrDtl01Div > li:nth-child(12) > a').get_attribute("href"))
sleep(1)

checklist_PP=driver.find_elements(By.CSS_SELECTOR,'[id^="chkAttrVal"][value*="플라스틱"]')
checklist_메쉬=driver.find_elements(By.CSS_SELECTOR,'[id^="chkAttrVal"][value*="메쉬"]')
checklist_가죽=driver.find_elements(By.CSS_SELECTOR,'[id^="chkAttrVal"][value*="가죽"]')

for check_메쉬 in checklist_메쉬:
    if not check_메쉬 in checklist_가죽:
        check_메쉬.click()           

for check_PP in checklist_PP:
    if not check_PP in checklist_가죽:
        check_PP.click()       

Execute_script(driver.find_element(By.CSS_SELECTOR,'#btnAttrInit > a:nth-child(2)').get_attribute("onclick"))

Execute_script(driver.find_element(By.CSS_SELECTOR,'#attrDtl01Div > li:nth-child(8) > a').get_attribute("href"))
sleep(1)

driver.find_element(By.CSS_SELECTOR,'[id^="chkAttrVal"][value^="유"]').click()
Execute_script(driver.find_element(By.CSS_SELECTOR,'#btnAttrInit > a:nth-child(2)').get_attribute("onclick"))

Execute_script(driver.find_element(By.CSS_SELECTOR,'#attrDtl01Div > li:nth-child(14) > a').get_attribute("href"))
sleep(1)

driver.find_element(By.CSS_SELECTOR,'[id^="chkAttrVal"][value^="유"]').click()
Execute_script(driver.find_element(By.CSS_SELECTOR,'#btnAttrInit > a:nth-child(2)').get_attribute("onclick"))

driver.find_element(By.CSS_SELECTOR,'#priorPrdCrtfcCheck').click()
sleep(1)

Select(driver.find_element(By.CSS_SELECTOR,'#pageSizeChkBox')).select_by_value("100")
sleep(1)

Execute_script(driver.find_element(By.CSS_SELECTOR,'#srchItemOpt > div > a').get_attribute("href"))
sleep(1)

img_list=driver.find_elements(By.CSS_SELECTOR,'tr > td > a > img')
for i,img in enumerate(img_list):
    imgurl=img.get_attribute('src')
    imgsrc=requests.get(imgurl,stream=True)
    with open(os.path.join(download_dir,f'{i}.png'),'wb') as f:
        shutil.copyfileobj(imgsrc.raw,f)

item_list=driver.find_elements(By.CSS_SELECTOR,'td > ul > li > a[href^="javascript:toSMPPGoodsDtlInfo("]')
item_list=list(map(lambda a:a.get_attribute("href"),item_list))

excel_filename=r"C:\Users\k\Documents\Hangeul Sample\webcrawling_2\chair_list.xlsx"
workbook=Workbook()
workbook.save(excel_filename)

with pd.ExcelWriter(excel_filename,engine='openpyxl',mode='a',if_sheet_exists='overlay') as writer:
    # Single sheet-Workbook, Multiple sheets-Excelwriter
    # mode='a': append mode
    
    for j, item in enumerate(item_list):
        Execute_script(item)
        sleep(2)

        spec=pd.read_html(driver.page_source,index_col=0)[0].transpose()
        if j==0:
            spec.columns=map(lambda a:a.replace(':',''),spec.columns)
            spec.to_excel(writer,startrow=0,sheet_name='Sheet',index=False)
        else:
            spec.to_excel(writer,startrow=writer.sheets['Sheet'].max_row,sheet_name='Sheet',index=False,header=False)            
         
driver.close()

spec_table=pd.read_excel(excel_filename)

hwp=win32.Dispatch("HWPFrame.HwpObject")
hwp.XHwpWindows.Item(0).Visible=True
hwp.RegisterModule("FilePathCheckDLL","SecurityModule")
hwp.Open(r"Downloads\셀주소찾기&amp;찾아가기.hwp",'','')
hwp.Open(r"C:\Users\k\Documents\Hangeul Sample\webcrawling_2\chair_list.hwp",'','')

def Numbering(i):
	hwp.HAction.GetDefault("ComposeChars", hwp.HParameterSet.HChCompose.HSet)
	hwp.HParameterSet.HChCompose.CharShapes.CircleCharShape.FontTypeHangul = hwp.FontType("TTF")		
	hwp.HParameterSet.HChCompose.CharShapes.CircleCharShape.FontTypeOther = hwp.FontType("TTF")
	hwp.HParameterSet.HChCompose.CharShapes.CircleCharShape.FontTypeSymbol = hwp.FontType("TTF")
	hwp.HParameterSet.HChCompose.CharShapes.CircleCharShape.FontTypeUser = hwp.FontType("TTF")
	hwp.HParameterSet.HChCompose.CircleType = 1
	hwp.HParameterSet.HChCompose.CheckCompose = 0
	hwp.HParameterSet.HChCompose.Chars = f"{i}"	
	hwp.HAction.Execute("ComposeChars", hwp.HParameterSet.HChCompose.HSet)

Act=hwp.CreateAction("TablePropertyDialog")
Set=Act.CreateSet()
# : Table
Pset=Set.CreateItemSet("ShapeTableCell","Cell")
# ParameterSet CreateItemSet(string itemid,string setid): Cell
Act.GetDefault(Set)
# Initialize set(Set)
w=(Pset.Item("Width")*0.95)/7200*25.4        
# Get length in HwpUnit (1 mm=283.465HwpUnit, 1 in=7200HwpUnit)
Act.Execute(Set)

# Another method by copying image to the cliboard

for idx in range(len(spec_table)):
    with Image.open(os.path.join(download_dir,f'{idx}.png')) as img:        
        h=w*(img.height/img.width)
        hwp.InsertPicture(img.filename,True,1,False,False,0,w,h)

    if idx%2==0:
        hwp.Run("TableAppendRow")
    else:
        hwp.Run("TableLowerCell")
        # hwp.Run("MoveDown")
        
    Numbering(idx+1)
    hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
    hwp.HParameterSet.HInsertText.Text = "   {}\r\n{}\r\n{}".format(str(spec_table.loc[idx,'업체명 ']).split("[")[0].strip(),
                                                         str(spec_table.loc[idx,'규격명 ']).split(",")[3].split("(")[0].strip(),
                                                         str(spec_table.loc[idx,'가격 ']))
    # '\r\n': new line feed
    # column name can be used as an attribute, But should be aware of whitespaces (e.g. spec_table.규격명)
    hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
    
    hwp.Run("TableRightCellAppend")
    if idx%2==0:
        hwp.Run("TableUpperCell")
        # hwp.Run("MoveUp")

## Exercise 6. Attach documents
# Ctrl+O
import win32com.client as win32

hwp=win32.gencache.EnsureDispatch("HWPFrame.HwpObject")
hwp.XHwpWindows.Item(0).Visible=True
hwp.RegisterModule("FilePathCheckDLL","SecurityModule")
hwp.Run("FileNew")

hwp.HAction.GetDefault("InsertFile", hwp.HParameterSet.HInsertFile.HSet)
hwp.HParameterSet.HInsertFile.FileName = "C:\\Users\\k\\Documents\\Hangeul Sample\\certificate_of_award\\h25_certificate_of_award.hwp"
hwp.HParameterSet.HInsertFile.KeepSection = 1
hwp.HParameterSet.HInsertFile.KeepCharshape = 1
hwp.HParameterSet.HInsertFile.KeepParashape = 1
hwp.HParameterSet.HInsertFile.KeepStyle = 1
hwp.HAction.Execute("InsertFile", hwp.HParameterSet.HInsertFile.HSet)

## Exercise 7. Compare two documents
import difflib
# library that helps compare two different strings
from time import sleep
import pyperclip as cb
# control clipboard
import win32com.client as win32

hwp=win32.Dispatch("HWPFrame.HwpObject")
hwp.XHwpWindows.Item(0).Visible=True
hwp.RegisterModule("FilePathCheckDLL","SecurityModule")
hwp.Open(r"Downloads\셀주소찾기&amp;찾아가기.hwp",'','')

def fontcolor(Color):
    hwp.HAction.Run('TableCellBlock')
    hwp.HAction.Run(f'CharShapeTextColor{Color.capitalize()}')
    hwp.HAction.Run('Cancel')

if __name__=='__main__':
    # Execute code below when execute the file externally, but do not execute when import this file to the other file
    hwp=win32.Dispatch("HWPFrame.HwpObject")
    hwp.XHwpWindows.Item(0).Visible=True
    hwp.RegisterModule("FilePathCheckDLL","SecurityModule")

    hwp.Run('FileNew')
    # Open new window
    hwp.Open(r"C:\Users\k\Documents\Hangeul Sample\compare_documents\별헤는밤.hwp",'','')
    hwp.Run('FileNew')
    hwp.Open(r"C:\Users\k\Documents\Hangeul Sample\compare_documents\별헤는밤_조작.hwp",'','')
    hwp.Run('FileNew')
    hwp.Open(r"C:\Users\k\Documents\Hangeul Sample\compare_documents\비교표.hwp",'','')

    origin=hwp.XHwpDocuments.Item(1)
    copy=hwp.XHwpDocuments.Item(2)
    diff=hwp.XHwpDocuments.Item(3)

    origin.SetActive_xHwpDocument()
    # Activate the origin window
    hwp.InitScan(None,None,None,None,None,None)
    original_full_text=''
    stop_signal=1
    while stop_signal:
        signal,text=hwp.GetText()
        if signal==1:   # End of the doc
            break
        original_full_text+=text
    hwp.ReleaseScan()
    original_full_text=original_full_text.split('\r\n')[:-1]

    copy.SetActive_xHwpDocument()
    # Activate the origin window
    hwp.InitScan(None,None,None,None,None,None)
    copy_full_text=''
    stop_signal=1
    while stop_signal:
        signal,text=hwp.GetText()
        if signal==1:   # End of the doc
            break
        copy_full_text+=text
    hwp.ReleaseScan()
    copy_full_text=copy_full_text.split('\r\n')[:-1]

    diff.SetActive_XHwpDocument()
    for original_statement in original_full_text:
        cb.copy(original_statement)
        hwp.Run('Paste')
        hwp.Run('TableRightCellAppend')
        coupled_dict=dict()
        for copy_statement in copy_full_text:
            coupled_dict[difflib.SequenceMatcher(None,original_statement.split(' ',1)[1],
                                                copy_statement.split(' ',1)[1]).ratio()]=copy_statement
        max_ratio=max(k for k, v in coupled_dict.items())
        cb.copy(coupled_dict[max_ratio].strip())
        hwp.Run('Paste')
        if max_ratio<1.0:   
            # If not a complete match
            fontcolor('red')
        else:
            fontcolor('black')

        hwp.Run('TableRightCellAppend')

    hwp.SaveAs(r"C:\Users\k\Documents\Hangeul Sample\compare_documents\최종.hwp",'','')
    hwp.Quit()

## Exercise 8. Navigate to the specific cell in the table
# hwp.MoveToField(fieldname,select=true/false,start=true/false)
# select: TableCellBlock, start: beginning/end of the fieldtext
# hwp.PutFieldText(Fieldname,text)

hwp=win32.Dispatch("HWPFrame.HwpObject")
hwp.XHwpWindows.Item(0).Visible=True
hwp.RegisterModule("FilePathCheckDLL","SecurityModule")
hwp.Open(r"C:\Users\k\Downloads\셀주소찾기&amp;찾아가기.hwp",'','')
# To get successful result from the method below, the carrot must be in the table
def SetTableCellAddr(addr):
    init_addr=hwp.KeyIndicator()[-1][1:].split(")")[0]
    if not hwp.CellShape:
        raise AttributeError('Carrot is not in the table')
    if addr==init_addr:
        return
    hwp.Run('CloseEx')
    hwp.FindCtrl()
    hwp.Run('ShapeObjTableSelCell')
    if addr=='A1':
        return
    while True:
        current_addr=hwp.KeyIndicator()[-1][1:].split(")")[0]
        hwp.Run("TableRightCell")
        if current_addr==hwp.KeyIndicator()[-1][1:].split(")")[0]:
            SetTableCellAddr(init_addr)
            hwp.Run('Cancel')
            raise AttributeError('Input data is out of range for the table')
        if addr==hwp.KeyIndicator()[-1][1:].split(")")[0]:
            return
    
# Gather multiple raw documents into excel
import win32com.client as win32
from openpyxl import Workbook

hwp=win32.gencache.EnsureDispatch("HWPFrame.HwpObject")
hwp.XHwpWindows.Item(0).Visible=True
hwp.RegisterModule("FilePathCheckDLL","SecurityModule")
hwp.Open(r"C:\Users\k\Downloads\성과목표지표계획서(19년점검결과)_산업부_에너지인력양성에특N전력기금통합.hwp",'','')

def get_text():
    hwp.InitScan(None,0x00ff,None,None,None,None)
    state,total_text=hwp.GetText()
    while state not in [0,1]:
        state,text=hwp.GetText()
        total_text+=text
    hwp.ReleaseScan()
    return total_text

hwp.HAction.GetDefault("RepeatFind", hwp.HParameterSet.HFindReplace.HSet);
hwp.HParameterSet.HFindReplace.FindString = "2. 성과목표 및 지표 총괄표"
hwp.HParameterSet.HFindReplace.Direction=2 
hwp.HParameterSet.HFindReplace.WholeWordOnly = 1
hwp.HParameterSet.HFindReplace.SeveralWords = 1,
hwp.HParameterSet.HFindReplace.IgnoreMessage = 1 
hwp.HParameterSet.HFindReplace.FindType = 1
hwp.HAction.Execute("RepeatFind", hwp.HParameterSet.HFindReplace.HSet)

hwp.MovePos(10)
hwp.FindCtrl()
hwp.Run('ShapeObjTableSelCell')

# ctrl=hwp.HeadCtrl
# while ctrl:
#     if ctrl.CtrlID=="tbl":
#         print('Table!')
#         hwp.SetPosBySet(ctrl.GetAnchorPos(0))
#         break
#     else:
#         ctrl=ctrl.Next

# contents=[]
# contents.append(get_text())
# while hwp.HAction.Run("TableRightCell"):
#     # True when the action is executed, False when the action can't be executed (last cell)
#     contents.append(get_text())

table_contents=dict()
table_contents[hwp.KeyIndicator()[-1][1:].split(")")[0]]=get_text()
while hwp.HAction.Run("TableRightCell"):
    # True when the action is executed, False when the action can't be executed (last cell)
    table_contents[hwp.KeyIndicator()[-1][1:].split(")")[0]]=get_text()

wb=Workbook()
ws=wb.active

for cell_addr, cell_content in table_contents.items():
    ws[cell_addr]=cell_content

wb.save(r"C:\Users\k\Documents\Hangeul Sample\test_using_Hangeul.xlsx")

# Exercise 9. Split and save the document
import os
from time import sleep
from tkinter.filedialog import askopenfilename
import win32com.client as win32

class Hwp:
    def __init__(self):
        self.hwp=win32.gencache.EnsureDispatch('HWPFrame.HwpObject')

    def __del__(self):
        self.hwp.Clear(option=1)
        self.hwp.Quit()

    def open_file(self,filename,view=False):
        self.name=filename
        self.hwp.RegisterModule("FilePathCheckDLL","SecurityModule")
        if view==True:
            self.hwp.Run('FileNew')
        self.hwp.Open(self.name)

    def split_save(self,name):
        self.hwp.MovePos(0)
        self.pagecount=self.hwp.PageCount
        hwp_docs=self.hwp.XHwpDocuments
        # To access multiple Hangeul windows

        target_folder=os.path.join(os.environ['USERPROFILE'],'desktop','result')
        try:
            os.mkdir(target_folder)
        except FileExistsError:
            print('바탕화면에 result 폴더가 이미 생성되어 있습니다')

        for i in range(self.pagecount):
            hwp_docs.Item(0).SetActive_XHwpDocument()
            sleep(1)
            self.hwp.Run('CopyPage')
            # Copy contents and style of the page at once 
            sleep(1)
            hwp.docs.Add(isTab=True)
            hwp.docs.Item(1).SetActive_XHwpDocument()
            self.hwp.Run('Paste')
            self.hwp.SaveAs(
                os.path.join(target_folder,name.rsplit('/')[-1].rsplit('.')[0]+'_'+str(i+1)+'.hwp')
            )            
            self.hwp.Run('FileClose')
            self.hwp.Run('MovePageDown')
            print(f'{i+1}/{self.pagecount}')

    def quit(self):
        self.hwp.Quit()

def main():
    name=askopenfilename(initialdir=os.path.join(os.environ['USERPROFILE'],'desktop'),
                         filetypes=(('아래아한글 파일','*.hwp'),('모든 파일','*.*')),
                         title='HWP 파일을 선택하세요')
    hwp=Hwp()
    hwp.open_file(name)
    hwp.split_save(name)
    hwp.quit()

if __name__=='__main__':
    main()

    